
from kivy.uix.screenmanager import SlideTransition
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.button import Button

from kivy.properties import StringProperty
from kivy.clock import Clock

from Modulos.Singleton.Perfil import Singleton_Perfil
from kivy.factory import Factory

from Modulos.BaseDatos.Conexion import Escritura_Eventos_DB

from datetime import datetime, time
import re

import string

import os
from openpyxl import load_workbook

try:
    from android.storage import primary_external_storage_path
    # Obtener la ruta del almacenamiento externo (típicamente /storage/emulated/0)
    DEFAULT_PATH = primary_external_storage_path() 
except ImportError:
    # Si no estamos en Android (estamos en Windows/Linux/macOS), usamos el directorio de usuario
    DEFAULT_PATH = os.path.expanduser('~')


class Input_Base(BoxLayout):
    letras_permitido = StringProperty('')
    t_text = StringProperty('Sin Texto')
    caracteres_extra = StringProperty(' ñÑÁÉÍÓÚáéíóúäëïöüÄËÏÖÜ,.')

    def __init__(self, **kwargs):
        super(Input_Base, self).__init__(**kwargs)
        self.letras_permitido = string.ascii_letters + string.digits
    
    def on_kv_post(self, base_widget):
        text_input = self.ids.get('usuario')
        if text_input:
            text_input.bind(text=self.filtrar_caracteres)
        
    def filtrar_caracteres(self, instance, value):
        if not value:
            return
        nuevo_texto_filtrado = ''
        
        for char in value:
            if char in self.letras_permitido or char in self.caracteres_extra:
                nuevo_texto_filtrado += char
        
        if value != nuevo_texto_filtrado:
            instance.unbind(text=self.filtrar_caracteres)
            instance.text = nuevo_texto_filtrado
            instance.bind(text=self.filtrar_caracteres)
            instance.cursor = (len(nuevo_texto_filtrado), 0)


class Input_CreacionEventos_Nombre(Input_Base):
    pass


class Input_CreacionEventos_numero(Input_Base):
    def __init__(self, **kwargs):
        super(Input_CreacionEventos_numero, self).__init__(**kwargs)
        self.letras_permitido = '0123456789-.'
        self.caracteres_extra = ''


class Input_CreacionEventos_Descripcion(Input_Base):
    pass


from threading import Thread
from Modulos.BaseDatos.Ubicacion import LocationService

class Input_CreacionEventos_Ubicacion(Input_Base):
    def on_kv_post(self, base_widget):
        text_input = self.ids.get('usuario')
        if text_input:
            text_input.bind(text=self.filtrar_caracteres)
            text_input.bind(focus=self.on_focus)

    def on_focus(self, instance, value):
        if not value: # Lost focus
            texto = instance.text.strip()
            if texto:
                # Run geocoding in a separate thread
                Thread(target=self.buscar_coordenadas, args=(texto,)).start()

    def buscar_coordenadas(self, direccion):
        service = LocationService()
        resultado = service.obtener_coordenadas_de_lugar(direccion)
        
        if resultado['latitud'] and resultado['longitud']:
            # Update UI on main thread
            Clock.schedule_once(lambda dt: self.actualizar_campos(resultado), 0)
        else:
             print(f"Geocoding failed: {resultado['mensaje']}")

    def actualizar_campos(self, resultado):
        # Traverse up to find the Layout_CreacionEventos (or a widget containing Latitud/Longitud)
        root = self
        found = False
        # Limit traversal to avoid infinite loops or going too far
        for _ in range(10): 
            if hasattr(root, 'ids') and 'Latitud' in root.ids and 'Longitud' in root.ids:
                found = True
                break
            if not root.parent:
                break
            root = root.parent
        
        if found:
            try:
                # Now that we changed the KV id to 'usuario', this should work
                root.ids.Latitud.ids.usuario.text = str(resultado['latitud'])
                root.ids.Longitud.ids.usuario.text = str(resultado['longitud'])
            except Exception as e:
                print(f"Error updating fields: {e}")
        else:
            print("Could not find Latitud/Longitud fields in parent hierarchy.")


class Input_CreacionEventos_Fecha(BoxLayout):
    # Declarar la propiedad para que sea accesible y reactiva
    letras_permitido = StringProperty('')
    #bg_icon = StringProperty('0.7,0.1,0.7,1') # Ejemplo: si necesitas otras propiedades
    #bg_text = StringProperty('1,1,1,1')
    t_text = StringProperty('Sin Texto')
    #password = StringProperty('False') # Kivy trata los valores del KV como strings

    def __init__(self, **kwargs):
        super(Input_CreacionEventos_Fecha, self).__init__(**kwargs)
        self.letras_permitido = '0123456789'
    
    def on_kv_post(self, base_widget):
        text_input = self.ids.get('usuario')
        
        if text_input:
            text_input.bind(text=self.filtrar_caracteres)
            text_input.bind(focus=self.on_focus)

    def on_focus(self, instance, value):
        """
        Maneja el evento de foco.
        value es True si tiene foco, False si lo pierde.
        """
        if value:
            # Al ganar foco: Quitar formato, dejar solo números
            texto_limpio = "".join([c for c in instance.text if c in self.letras_permitido])
            instance.text = texto_limpio
        else:
            # Al perder foco: Aplicar formato
            self.aplicar_formato(instance)

    def aplicar_formato(self, instance):
        """Aplica el formato de fecha (DD-MM-YYYY) al texto actual."""
        value = instance.text
        caracteres_puros = [char for char in value if char in self.letras_permitido]
        
        # Validar fecha (solo para color)
        es_fecha_valida = True
        try:
            temp_str = "".join(caracteres_puros)
            if len(temp_str) >= 1:
                day = int(temp_str[:2]) if len(temp_str) >= 2 else 1
                month = int(temp_str[2:4]) if len(temp_str) >= 4 else 1
                year = int(temp_str[4:8]) if len(temp_str) >= 8 else 2000
                datetime(day=day, month=month, year=year)
        except:
            es_fecha_valida = False
            
        if es_fecha_valida:
            self.bg_text = 1, 1, 1, 1
        else:
            self.bg_text = 0.6, 0.2, 0.2, 1

        # Formatear
        nuevo_texto_con_formato = list(caracteres_puros[:8]) # Max 8 dígitos
        
        indices_separador = [4, 2] 
        for i in indices_separador:
            if len(nuevo_texto_con_formato) > i:
                nuevo_texto_con_formato.insert(i, '-')
                
        texto_final = "".join(nuevo_texto_con_formato)
        
        # Actualizar texto sin disparar bind recursivo (aunque filtrar_caracteres ahora es simple)
        instance.unbind(text=self.filtrar_caracteres)
        instance.text = texto_final
        instance.bind(text=self.filtrar_caracteres)
            
            
    def _posicionar_cursor(self, dt, text_input, new_cursor_x, cursor_fila_original):
        """
        ¡Este método debe estar aquí!
        """
        # ... (Contenido del método _posicionar_cursor)
        text_input.bind(text=self.filtrar_caracteres)
        text_input.cursor = (new_cursor_x, cursor_fila_original)
        
        
    def filtrar_caracteres(self, instance, value):
        """
        Solo filtra caracteres no permitidos y longitud mientras se escribe.
        El formato se aplica en on_focus (al perder foco).
        """
        if not value:
            return

        # Si tiene foco, solo permitimos números y limitamos longitud (8 dígitos)
        if instance.focus:
            caracteres_puros = "".join([char for char in value if char in self.letras_permitido])
            texto_final = caracteres_puros[:8]
            
            if value != texto_final:
                instance.unbind(text=self.filtrar_caracteres)
                
                # Guardar posición relativa del cursor
                cursor_index = instance.cursor_index()
                # Calcular nuevo cursor (simple, ya que solo borramos chars inválidos)
                # Si borramos algo antes del cursor, retrocedemos.
                # Para simplificar: cursor al final si cambia mucho, o intentar mantener.
                # Kivy maneja bien el cursor si solo reemplazamos texto.
                
                instance.text = texto_final
                instance.bind(text=self.filtrar_caracteres)
                
                # Re-posicionar cursor al final si se cortó texto, o dejarlo si es edición interna
                # instance.cursor = (len(texto_final), 0) 
        else:
            # Si no tiene foco (ej: seteo programático), aplicamos formato directo
            # Esto puede pasar si se setea texto desde código
            pass


class Input_CreacionEventos_Hora(BoxLayout):
    # Declarar la propiedad para que sea accesible y reactiva
    letras_permitido = StringProperty('')
    t_text = StringProperty('Sin Texto')

    def __init__(self, **kwargs):
        super(Input_CreacionEventos_Hora, self).__init__(**kwargs)
        self.letras_permitido = '0123456789'
    
    def on_kv_post(self, base_widget):
        """Se llama después de cargar el KV para enlazar los TextInputs."""
        text_input = self.ids.get('usuario')
        
        if text_input:
            text_input.bind(text=self.filtrar_caracteres)
            text_input.bind(focus=self.on_focus)
            
    def on_focus(self, instance, value):
        if value:
            # Ganó foco: Quitar formato
            texto_limpio = "".join([c for c in instance.text if c in self.letras_permitido])
            instance.text = texto_limpio
        else:
            # Perdió foco: Aplicar formato
            self.aplicar_formato(instance)

    def aplicar_formato(self, instance):
        value = instance.text
        caracteres_puros = [char for char in value if char in self.letras_permitido]
        
        # Validar hora (HHMM)
        es_hora_valida = True
        try:
            temp_str = "".join(caracteres_puros)
            # Auto-fill: Si solo hay 1-2 dígitos, asumir que son horas y rellenar minutos con 00
            if len(temp_str) == 1 or len(temp_str) == 2:
                temp_str = temp_str + "00"
            elif len(temp_str) == 3:
                temp_str = "0" + temp_str
            
            if len(temp_str) >= 4:
                hour = int(temp_str[:2])
                minute = int(temp_str[2:4])
                time(hour=hour, minute=minute)
        except:
            es_hora_valida = False
            
        if es_hora_valida:
            self.bg_text = 1, 1, 1, 1
        else:
            self.bg_text = 0.6, 0.2, 0.2, 1

        # Formatear HH:MM
        nuevo_texto_con_formato = list(caracteres_puros[:4]) # Max 4 dígitos
        
        # Auto-fill logic
        if len(nuevo_texto_con_formato) == 1 or len(nuevo_texto_con_formato) == 2:
            nuevo_texto_con_formato.extend(['0', '0'])
        elif len(nuevo_texto_con_formato) == 3:
            nuevo_texto_con_formato.insert(0, '0')
        
        if len(nuevo_texto_con_formato) > 2:
            nuevo_texto_con_formato.insert(2, ':')
                
        texto_final = "".join(nuevo_texto_con_formato)
        
        instance.unbind(text=self.filtrar_caracteres)
        instance.text = texto_final
        instance.bind(text=self.filtrar_caracteres)

    def filtrar_caracteres(self, instance, value):
        if not value:
            return

        if instance.focus:
            # Solo permitir números y limitar longitud (4 dígitos)
            caracteres_puros = "".join([char for char in value if char in self.letras_permitido])
            texto_final = caracteres_puros[:4]
            
            if value != texto_final:
                instance.unbind(text=self.filtrar_caracteres)
                instance.text = texto_final
                instance.bind(text=self.filtrar_caracteres)
        else:
            pass  


class Input_CreacionEventos_dinero(BoxLayout):
    # Declarar la propiedad para que sea accesible y reactiva
    letras_permitido = StringProperty('')
    #bg_icon = StringProperty('0.7,0.1,0.7,1') # Ejemplo: si necesitas otras propiedades
    #bg_text = StringProperty('1,1,1,1')
    t_text = StringProperty('Sin Texto')
    #password = StringProperty('False') # Kivy trata los valores del KV como strings

    def __init__(self, **kwargs):
        super(Input_CreacionEventos_dinero, self).__init__(**kwargs)
        self.letras_permitido = '0123456789'
    
    def on_kv_post(self, base_widget):
        text_input = self.ids.get('dinero')
        
        if text_input:
            text_input.bind(text=self.filtrar_caracteres)
            text_input.bind(focus=self.on_focus)

    def on_focus(self, instance, value):
        if value:
            # Ganó foco: Quitar formato
            texto_limpio = "".join([c for c in instance.text if c in self.letras_permitido])
            instance.text = texto_limpio
        else:
            # Perdió foco: Aplicar formato
            self.aplicar_formato(instance)

    def aplicar_formato(self, instance):
        value = instance.text
        # Limpiar y dejar solo dígitos
        texto_limpio = [v for v in value if v in self.letras_permitido]
        # Limitar a 6 dígitos (según tu lógica original <=5 index, o sea 6 chars)
        texto_limpio = texto_limpio[:6] 
        
        if not texto_limpio:
            return

        # Lógica de puntos (miles)
        # Invertimos, agrupamos de a 3, unimos con punto, invertimos de nuevo
        # O usamos tu lógica de lista_puntos
        
        # Tu lógica original adaptada:
        lista_puntos = list(-n for n in range(3, len(texto_limpio), 3))
        lista_puntos.sort() # Ordenar para insertar desde el final (-3, -6...)
        
        for n in lista_puntos:
            if len(texto_limpio) > (n*-1):
                texto_limpio.insert(n, '.')
        
        texto_limpio.insert(0, 'CLP $')
        texto_formateado = "".join(texto_limpio)
        
        instance.unbind(text=self.filtrar_caracteres)
        instance.text = texto_formateado
        instance.bind(text=self.filtrar_caracteres)
        self.bg_text = 1, 1, 1, 1
        
    def _posicionar_cursor(self, dt, text_input, new_cursor_x, cursor_fila_original):
        """
        ¡Este método debe estar aquí!
        """
        # ... (Contenido del método _posicionar_cursor)
        text_input.bind(text=self.filtrar_caracteres)
        text_input.cursor = (new_cursor_x, cursor_fila_original)    
            
    def filtrar_caracteres(self, instance, value):
        if not value:
            return

        if instance.focus:
            # Solo permitir números y limitar longitud
            caracteres_puros = "".join([char for char in value if char in self.letras_permitido])
            texto_final = caracteres_puros[:6] # Max 6 dígitos
            
            if value != texto_final:
                instance.unbind(text=self.filtrar_caracteres)
                instance.text = texto_final
                instance.bind(text=self.filtrar_caracteres)
        else:
            pass


class Input_CreacionEventos_rut(BoxLayout):
    letras_permitido = StringProperty('')
    t_text = StringProperty('Sin Texto')

    def __init__(self, **kwargs):
        super(Input_CreacionEventos_rut, self).__init__(**kwargs)
        self.letras_permitido = '0123456789Kk'
    
    def on_kv_post(self, base_widget):
        text_input = self.ids.get('rut')
        if text_input:
            text_input.bind(text=self.filtrar_caracteres)
            text_input.bind(focus=self.on_focus)

    def on_focus(self, instance, value):
        if value:
            # Ganó foco: Quitar formato
            texto_limpio = "".join([c for c in instance.text if c in self.letras_permitido])
            instance.unbind(text=self.filtrar_caracteres)
            instance.text = texto_limpio
            instance.bind(text=self.filtrar_caracteres)
        else:
            # Perdió foco: Aplicar formato
            self.aplicar_formato(instance)

    def aplicar_formato(self, instance):
        value = instance.text
        caracteres_puros = [char for char in value if char in self.letras_permitido]
        
        if len(caracteres_puros) < 2:
            return
        
        digito_verificador = caracteres_puros[-1].upper() if caracteres_puros[-1].lower() == 'k' else caracteres_puros[-1]
        numeros = caracteres_puros[:-1]
        
        # Limitar a 8 dígitos (99.999.999)
        numeros = numeros[:8]
        
        numeros_str = ''.join(numeros)
        numeros_reversed = numeros_str[::-1]
        grupos = [numeros_reversed[i:i+3] for i in range(0, len(numeros_reversed), 3)]
        numeros_formateados = '.'.join(grupos)[::-1]
        
        texto_final = f"{numeros_formateados}-{digito_verificador}"
        
        # Validar
        if self.verificacion_rut(texto_final):
            self.bg_text = 1, 1, 1, 1
        else:
            self.bg_text = 1, 0, 0, 1
        
        instance.unbind(text=self.filtrar_caracteres)
        instance.text = texto_final
        instance.bind(text=self.filtrar_caracteres)

    def filtrar_caracteres(self, instance, value):
        if not value:
            return

        if instance.focus:
            # Solo permitir números, K y k, limitar longitud (max 9: 8 dígitos + 1 DV)
            caracteres_puros = "".join([char for char in value if char in self.letras_permitido])
            texto_final = caracteres_puros[:9]
            
            if value != texto_final:
                instance.unbind(text=self.filtrar_caracteres)
                instance.text = texto_final
                instance.bind(text=self.filtrar_caracteres)
        else:
            pass
    
    def verificacion_rut(self, rut_completo):
        rut_limpio = re.sub(r'[.-]', '', rut_completo).upper()
        if not rut_limpio or len(rut_limpio) < 2:
            return False
            
        cuerpo = rut_limpio[:-1]
        dv = rut_limpio[-1]
        
        if not cuerpo.isdigit():
            return False
            
        # VALIDACION NUMEROS REPETIDOS (Requested by user)
        if len(set(cuerpo)) == 1:
            return False

        rut_invertido = cuerpo[::-1]
        multiplicador = 2
        suma = 0
        for digito in rut_invertido:
            suma += int(digito) * multiplicador
            multiplicador += 1
            if multiplicador > 7:
                multiplicador = 2
        resto = suma % 11
        dv_calculado = 11 - resto
        if dv_calculado == 11:
            dv_esperado = '0'
        elif dv_calculado == 10:
            dv_esperado = 'K'
        else:
            dv_esperado = str(dv_calculado)
            
        return dv == dv_esperado


class FileSelectPopup(Popup):
    def __init__(self, select_callback, filters, **kwargs):
        super().__init__(**kwargs)
        self.select_callback = select_callback
        self.title = 'Seleccionar Archivo de Invitados'
        self.size_hint = (0.9, 0.9)

        # Contenedor principal del popup
        content = BoxLayout(orientation='vertical', spacing=10)

        # 1. Selector de archivos
        self.filechooser = FileChooserListView(
            filters=filters,
            multiselect=False,
            path=DEFAULT_PATH
        )
        content.add_widget(self.filechooser)

        # 2. Botones de acción
        action_layout = BoxLayout(size_hint_y=None, height=50, spacing=10)
        
        btn_cancel = Button(text='Cancelar', size_hint_x=0.3)
        btn_cancel.bind(on_release=self.dismiss)
        action_layout.add_widget(btn_cancel)
        
        btn_select = Button(text='Seleccionar', size_hint_x=0.7)
        btn_select.bind(on_release=self.on_select)
        action_layout.add_widget(btn_select)

        content.add_widget(action_layout)
        self.content = content

    def on_select(self, instance):
        if self.filechooser.selection:
            # Obtener la ruta del archivo seleccionado
            selected_file = self.filechooser.selection[0]
            self.select_callback(selected_file)
            self.dismiss()
        else:
            print("No se ha seleccionado ningún archivo.")

class Tabla_Invitados(BoxLayout):
    identificador = []
    def __init__(self, **kwargs):
        super(Tabla_Invitados,self).__init__(**kwargs)
        
    def seleccionar_archivo(self, tipo):
        """Muestra el selector de archivos y define la función de manejo."""
        
        if tipo == 'xlsx':
            filters = ['*.xlsx']
            # Definir la función que manejará el archivo XLSX
            def on_file_selected(path):
                self.Importar_invitados(path)
        else:
            return

        popup = FileSelectPopup(
            select_callback=on_file_selected, 
            filters=filters
        )
        popup.open()
        
        
    def Agregar_invitado(self,rut="",fechas=None):
        invitado= Factory.Filas_ListaInvitados()
        invitado.accion= self.Agregar_fecha
        invitado.ids.rut_invitado.ids.rut.text= rut
        
        numero = 1
        
        while True:
            if numero not in self.identificador:
                invitado.id=numero
                self.identificador.append(numero)
                break
            else:
              numero +=1
              
        if fechas:
            # print(fechas)
            for fecha in fechas:
                # print(fecha)
                self.Agregar_fecha(invitado,fecha)
        
        self.ids.filas_invitados.add_widget(invitado)
        
    def Agregar_fecha(self,fila, listafechas=None):
        fecha= Factory.Fechas_ListaInvitados()
        # print(listafechas)
        if listafechas:
            # Asumimos formato "YYYY-MM-DD HH:MM" o "DD-MM-YYYY HH:MM"
            # Separamos fecha y hora
            try:
                inicio_parts = listafechas[0].split(' ')
                fecha.ids.fecha_inicio.ids.usuario.text = inicio_parts[0]
                if len(inicio_parts) > 1:
                    # Ahora solo seteamos el campo usuario con HHMM (sin :)
                    hora_str = inicio_parts[1].replace(':', '')
                    fecha.ids.hora_inicio.ids.usuario.text = hora_str
            except:
                pass # Manejo de errores simple
            
            try:
                fin_parts = listafechas[1].split(' ')
                fecha.ids.fecha_fin.ids.usuario.text = fin_parts[0]
                if len(fin_parts) > 1:
                    hora_str = fin_parts[1].replace(':', '')
                    fecha.ids.hora_fin.ids.usuario.text = hora_str
            except:
                pass
           
        
        numero = 1
        while True:
            if numero not in self.identificador:
                fecha.id=numero
                self.identificador.append(numero)
                break
            else:
              numero +=1
        # print(listafechas)
        fila.ids.fechas_invitado.add_widget(fecha)
        
    def Importar_invitados(self,ruta_archivo):
        try:
            # Reemplaza la simulación con la carga real de openpyxl
            wb = load_workbook(filename=ruta_archivo)
            sheet = wb.active
        except Exception as e:
            # print(ruta_archivo)
            print(f"Error al cargar el archivo XLSX: {e}")
            return
        
        
        self.ids.filas_invitados.clear_widgets()
        datos_importados = [('11.111.111-2','2024-07-01 10:00','2024-07-01 12:00')]
        
        datos_formateados = {}
        for i, dato in enumerate(sheet.iter_rows(min_row=2, values_only=True)):    #datos_importados:
            datos_formateados.setdefault(str(dato[0]), []).append((str(dato[1]), str(dato[2])))
        #print(datos_formateados)
        for rut, fechas in datos_formateados.items():
            #print(fechas)
            self.Agregar_invitado(rut, fechas)
        



class Layout_CreacionEventos(BoxLayout):
    def __init__(self, abrir_otra_pantalla, **kwargs):
        super(Layout_CreacionEventos,self).__init__(**kwargs)
        self.abrir_otra_pantalla= abrir_otra_pantalla
        #self.Desplegar_Visibilidad()
        
    def Abrir_MenuPrincipal(self):
        # Limpiar formulario antes de salir
        self.Limpiar_Campos()
        
        rol = Singleton_Perfil.get_instance().tipo_perfil
        if rol == 'Estandar':
            self.abrir_otra_pantalla("BA_Estandar", transition=SlideTransition(direction="right"))
        elif rol == 'Organizador':
            self.abrir_otra_pantalla("BB_Organizador", transition=SlideTransition(direction="right"))
        elif rol == 'Administrador':
            self.abrir_otra_pantalla("BC_Administrador", transition=SlideTransition(direction="right"))
        else:
            self.abrir_otra_pantalla("AA_Login", transition=SlideTransition(direction="right"))
    
    
    def Abrir_Login(self):
        self.abrir_otra_pantalla("AA_Login", transition= SlideTransition(direction="right"))
    
    def Limpiar_Campos(self):
        """
        Limpia todos los campos del formulario de creación de eventos.
        """
        print("🧹 Iniciando limpieza de campos...")
        
        # 1. Limpiar campos de texto básicos
        # Usamos try-except para evitar que un error detenga toda la limpieza
        try:
            self.ids.NombreEvento.ids.usuario.text = ''
            self.ids.Descripcion.ids.usuario.text = ''
            self.ids.Ubicacion.ids.usuario.text = ''
            self.ids.Latitud.ids.usuario.text = ''
            self.ids.Longitud.ids.usuario.text = ''
        except Exception as e:
            print(f"⚠️ Error limpiando campos de texto: {e}")

        # 2. Limpiar fechas y horas
        try:
            self.ids.FechaInicio.ids.usuario.text = ''
            self.ids.HoraInicio.ids.usuario.text = ''
            self.ids.FechaFin.ids.usuario.text = ''
            self.ids.HoraFin.ids.usuario.text = ''
        except Exception as e:
            print(f"⚠️ Error limpiando fechas/horas: {e}")
        
        # 3. Resetear spinners
        try:
            self.ids.acceso_spinner.text = 'Tipo de Acceso'
            self.ids.visibilidad_spinner.text = 'Tipo Evento'
        except Exception as e:
            print(f"⚠️ Error reseteando spinners: {e}")
        
        # 4. Limpiar campos de dinero
        try:
            self.ids.entrada_precio.ids.dinero.text = ''
            self.ids.preinscripcion_precio.ids.dinero.text = ''
        except Exception as e:
            print(f"⚠️ Error limpiando precios: {e}")
        
        # 5. Limpiar etiquetas (Lógica corregida)
        try:
            lista_etiquetas = self.ids.get('Lista_Etiquetas')
            if lista_etiquetas and hasattr(lista_etiquetas, 'ids') and 'Elementos' in lista_etiquetas.ids:
                contenedor_elementos = lista_etiquetas.ids.Elementos
                for child in contenedor_elementos.children:
                    if hasattr(child, 'ids') and 'seleccion' in child.ids:
                        child.ids.seleccion.active = False
        except Exception as e:
            print(f"⚠️ Error limpiando etiquetas: {e}")
        
        # 6. Eliminar tabla de invitados
        try:
            contenedor = self.ids.get('contenedor')
            if contenedor:
                # Buscar widget con id 'tabla_invitados'
                tabla_a_eliminar = None
                # Iteramos sobre una copia de la lista para poder modificarla
                for widget in contenedor.children[:]:
                    if hasattr(widget, 'id') and widget.id == 'tabla_invitados':
                        tabla_a_eliminar = widget
                        break
                
                if tabla_a_eliminar:
                    contenedor.remove_widget(tabla_a_eliminar)
                    print("✅ Tabla de invitados eliminada")
                    
                    # Limpiar referencia si existe
                    if hasattr(self, 'tabla_invitados_widget'):
                        self.tabla_invitados_widget = None
        except Exception as e:
            print(f"⚠️ Error eliminando tabla de invitados: {e}")
            
        print("✅ Formulario limpiado completamente")
        
    
    def Eliminar_TablaInvitados(self, widget, *args):
        """
        Elimina la tabla si la visibilidad es 'Publico'.
        El *args captura los argumentos extra del touch_up.
        """
        visibilidad = self.ids.visibilidad_spinner.text
        
        # 2. Aplicar la Condición: Solo eliminar si es 'Publico'
        if visibilidad == 'Publico': # Usamos "Publico" sin tilde, como lo tenías.
            # print(f"Visibilidad '{visibilidad}': Eliminando widget.")
            
            # Eliminar el widget del contenedor
            self.ids.contenedor.clear_widgets([widget])
            
            # 💥 CRÍTICO: Limpiar la referencia para indicar que ya no existe 💥
            self.tabla_invitados_widget = None
            
            # ❌ LÍNEA ELIMINADA: Esto causaba el NameError (tabla no definida)
            # y trababa el spinner al intentar reasignar su on_touch_up.
            # self.ids.visibilidad_spinner.on_touch_up = lambda *args: self.Eliminar_TablaInvitados(tabla, *args)
            
        else:
            # print(f"Visibilidad '{visibilidad}': No se permite la eliminación al tocar el widget.")
            return True # Indica que el evento fue manejado (detiene la propagación)
            
    def Agregar_TablaInvitados(self):
        visibilidad = self.ids.visibilidad_spinner.text
        
        # 1. Verificación de Visibilidad (Solo crear si es 'Privado')
        if visibilidad != 'Privado':
            # print(f"Visibilidad '{visibilidad}': No se permite agregar tabla.")
            return
            
        # 2. 🛑 VERIFICACIÓN DE EXISTENCIA CON PROPIEDAD 🛑
        # Nota: Si no estás usando ObjectProperty, este código debería usar el for loop seguro.
        # Por ahora, mantengo el for loop hasta que confirmes la definición de ObjectProperty.
        for widget in self.ids.contenedor.children:
            if hasattr(widget, 'id') and widget.id == 'tabla_invitados':
                # print("Advertencia: Ya existe una tabla de invitados. No se agregará otra.")
                return

        # Si la tabla ya existe, salimos
        # if self.tabla_invitados_widget:
        #     print("Advertencia: Ya existe una tabla de invitados. No se agregará otra.")
        #     return 
            
        # 3. Creación e inicialización
        tabla = Factory.Tabla_Invitados()
        tabla.id = 'tabla_invitados'
        
        # Guardar la referencia antes de añadir (Si usas ObjectProperty)
        # self.tabla_invitados_widget = tabla 
        
        # 4. Vinculación del Evento (en la tabla, no en el spinner)
        # Se vincula a la nueva instancia de 'tabla' para que se elimine al tocarla
        tabla.on_touch_up = lambda *args: self.Eliminar_TablaInvitados(tabla, *args)
        
        # ❌ LÍNEA ELIMINADA: Esto causaba la RecursionError y trababa el spinner.
        # self.ids.visibilidad_spinner.on_touch_down = self.Agregar_TablaInvitados()
        
        try:
            # 5. Añadir el widget
            self.ids.contenedor.add_widget(tabla)
            # print("Tabla de invitados agregada exitosamente.")
        except Exception as e:
            print(f"Error al añadir tabla: {e}. Limpiando referencia.")
            # Si falla al añadir, limpiamos la referencia (Si usas ObjectProperty)
            # self.tabla_invitados_widget = None 
            pass
        

    
    def Crear_Evento(self):
        # print('Entra')
        errores = {
            'campos vacios':0,
            'problemas en invitados':0
        }
        titulo = self.ids.NombreEvento.ids.usuario.text
        errores['campos vacios'] += 1 if titulo == '' else 0
        descripcion = self.ids.Descripcion.ids.usuario.text
        errores['campos vacios'] += 1 if descripcion == '' else 0
        ubicacion = self.ids.Ubicacion.ids.usuario.text
        errores['campos vacios'] += 1 if ubicacion == '' else 0
        latitud = self.ids.Latitud.ids.usuario.text
        errores['campos vacios'] += 1 if latitud == '' else 0
        longitud = self.ids.Longitud.ids.usuario.text
        errores['campos vacios'] += 1 if longitud == '' else 0
        
        
        
        Fecha_Inicio = ''
        try:
            Fecha_Inicio = datetime.strptime(self.ids.FechaInicio.ids.usuario.text,'%d-%m-%Y')
            
        except:
            errores['campos vacios'] += 1 # Si falla la conversión, el campo está mal o vacío.
            
        
        Fecha_Termino = ''
        try:
            Fecha_Termino = datetime.strptime(self.ids.FechaFin.ids.usuario.text,'%d-%m-%Y')
        except:
            errores['campos vacios'] += 1
            
        
        # Hora parsing - now using single field with HH:MM format
        try:
            hora_inicio_text = self.ids.HoraInicio.ids.usuario.text.replace(':', '')
            if len(hora_inicio_text) >= 3:
                if len(hora_inicio_text) == 3:
                    hora_inicio_text = '0' + hora_inicio_text
                Hora_Comienza = (int(hora_inicio_text[:2]), int(hora_inicio_text[2:4]))
            else:
                Hora_Comienza = (0, 0)
                errores['campos vacios'] += 1
        except:
            Hora_Comienza = (0, 0)
            errores['campos vacios'] += 1
        
        try:
            hora_fin_text = self.ids.HoraFin.ids.usuario.text.replace(':', '')
            if len(hora_fin_text) >= 3:
                if len(hora_fin_text) == 3:
                    hora_fin_text = '0' + hora_fin_text
                Hora_Finaliza = (int(hora_fin_text[:2]), int(hora_fin_text[2:4]))
            else:
                Hora_Finaliza = (0, 0)
                errores['campos vacios'] += 1
        except:
            Hora_Finaliza = (0, 0)
            errores['campos vacios'] += 1
        
        try:
            Fecha_Inicio = datetime(day=Fecha_Inicio.day,month=Fecha_Inicio.month,year=Fecha_Inicio.year, hour=Hora_Comienza[0],minute=Hora_Comienza[1])
        except:
            errores['campos vacios'] += 1
        try:
            Fecha_Termino = datetime(day=Fecha_Termino.day,month=Fecha_Termino.month,year=Fecha_Termino.year, hour=Hora_Finaliza[0],minute=Hora_Finaliza[1])
        except:
            errores['campos vacios'] += 1
        
        
        Etiquetas = []
        # -------------------------------------------------------------
        # !!! CORRECCIÓN DE LA LÓGICA DE ETIQUETAS !!!
        # Kivy GridLayout.children lista los widgets en orden inverso
        for etiqueta_widget in self.ids.Lista_Etiquetas.ids.Elementos.children[::-1]:
            # 'etiqueta_widget' es Etiquetas_Elementos
            
            # Si 'seleccion' está directamente dentro de etiqueta_widget (como ID):
            if hasattr(etiqueta_widget.ids, 'seleccion') and etiqueta_widget.ids.seleccion.active:
                # Si el texto está directamente en el widget:
                if hasattr(etiqueta_widget, 'text'):
                    Etiquetas.append(etiqueta_widget.text)
        # -------------------------------------------------------------
        
        errores['campos vacios'] += 1 if len(Etiquetas) == 0 else 0
        
        # ... (el resto de tu código sigue aquí sin cambios, usando Hora_Comienza y Hora_Finaliza)
        
        acceso = {}
        acceso_t = self.ids.acceso_spinner.text
        errores['campos vacios'] += 1 if acceso_t == '' else 0
        if acceso_t != '':
            acceso['Tipo'] = acceso_t
        
        acceso_v = None if acceso_t != 'Paga' else self.ids.entrada_precio.ids.dinero.text
        # Extraer todos los dígitos y unirlos (ej: "CLP $3.000" -> ['3', '000'] -> '3000')
        if acceso_v is not None:
            digitos = re.findall(r'[0-9]+', acceso_v)
            if digitos:
                #acceso['Tipo'] = acceso_t
                acceso['Valor'] = int(''.join(digitos))  # Unir todos los grupos de dígitos
            else:
                errores['campos vacios'] += 1
        else:
            if acceso_t == 'Paga':
                errores['campos vacios'] += 1
            
        
        visibilidad = {}
        visibilidad_t = self.ids.visibilidad_spinner.text
        errores['campos vacios'] += 1 if visibilidad_t == '' else 0
        if visibilidad_t != '':
            visibilidad['Tipo']= visibilidad_t
        
        visibilidad_v = None if visibilidad_t != 'Publico' else self.ids.preinscripcion_precio.ids.dinero.text
        # Extraer todos los dígitos y unirlos (ej: "CLP $15.000" -> ['15', '000'] -> '15000')
        if visibilidad_t == 'Publico':
            if visibilidad_v:
                digitos = re.findall(r'[0-9]+', visibilidad_v)
                if digitos:
                    visibilidad['Valor'] = int(''.join(digitos))  # Unir todos los grupos de dígitos
                else:
                    errores['campos vacios'] += 1
            else:
                errores['campos vacios'] += 1
        
        Invitados = {}
        if visibilidad_t == 'Privado':
            for widget in self.ids.contenedor.children:
                if hasattr(widget, 'id') and widget.id == 'tabla_invitados':
                    for i in widget.ids.filas_invitados.children:
                        periodos = {}
                        for n, f in enumerate(i.ids.fechas_invitado.children):
                            f_n=f.ids.fecha_inicio
                            f_t=f.ids.fecha_fin
                            fecha = {}
                            
                            f_n_t=f_n.ids.usuario.text
                            # print(f_n_t)
                            if f_n_t != '':
                                try:
                                    f_n_d_base = datetime.strptime(f_n_t,'%d-%m-%Y')
                                    # Obtener hora del campo usuario (formato HHMM o HH:MM)
                                    hora_inicio_text = f.ids.hora_inicio.ids.usuario.text.replace(':', '')
                                    if len(hora_inicio_text) >= 3:
                                        if len(hora_inicio_text) == 3:
                                            hora_inicio_text = '0' + hora_inicio_text
                                        h_i = int(hora_inicio_text[:2])
                                        m_i = int(hora_inicio_text[2:4])
                                    else:
                                        h_i, m_i = 0, 0
                                    
                                    f_n_d = f_n_d_base.replace(hour=h_i, minute=m_i)

                                    if f_n_d >= Fecha_Inicio and f_n_d <= Fecha_Termino:
                                        fecha['Comienza']=f_n_d.isoformat()
                                    else: 
                                        errores['problemas en invitados'] += 1
                                        fecha['Comienza']=Fecha_Inicio.isoformat()
                                except:
                                    # print("error detectado 1")
                                    errores['problemas en invitados'] += 1
                                    
                            else:
                                # Si no hay fecha, intentar usar la hora proporcionada con la fecha de inicio del evento
                                try:
                                    hora_inicio_text = f.ids.hora_inicio.ids.usuario.text.replace(':', '')
                                    if len(hora_inicio_text) >= 3:
                                        if len(hora_inicio_text) == 3:
                                            hora_inicio_text = '0' + hora_inicio_text
                                        h_i = int(hora_inicio_text[:2])
                                        m_i = int(hora_inicio_text[2:4])
                                        
                                        # Usar fecha de inicio del evento pero con la hora especificada
                                        f_n_d = Fecha_Inicio.replace(hour=h_i, minute=m_i)
                                        fecha['Comienza'] = f_n_d.isoformat()
                                    else:
                                        # Si no hay fecha ni hora, usar inicio del evento
                                        fecha['Comienza']=Fecha_Inicio.isoformat()
                                except:
                                    fecha['Comienza']=Fecha_Inicio.isoformat()
                            
                            f_t_t=f_t.ids.usuario.text
                            # print(f_t_t)
                            
                            if f_t_t !='':
                                try:
                                    f_t_d_base = datetime.strptime(f_t_t,'%d-%m-%Y')
                                    # Obtener hora del campo usuario (formato HHMM o HH:MM)
                                    hora_fin_text = f.ids.hora_fin.ids.usuario.text.replace(':', '')
                                    if len(hora_fin_text) >= 3:
                                        if len(hora_fin_text) == 3:
                                            hora_fin_text = '0' + hora_fin_text
                                        h_f = int(hora_fin_text[:2])
                                        m_f = int(hora_fin_text[2:4])
                                    else:
                                        h_f, m_f = 0, 0
                                    
                                    f_t_d = f_t_d_base.replace(hour=h_f, minute=m_f)

                                    if f_t_d <= Fecha_Termino and f_t_d >= fecha.get('Comienza', Fecha_Inicio):
                                        fecha['Termina']=f_t_d.isoformat()
                                    else: 
                                        errores['problemas en invitados'] += 1
                                        fecha['Termina']=Fecha_Termino.isoformat()
                                except:
                                    # print("error detectado 2")
                                    errores['problemas en invitados'] += 1
                            else:
                                # Si no hay fecha, intentar usar la hora proporcionada con la fecha de término del evento
                                try:
                                    hora_fin_text = f.ids.hora_fin.ids.usuario.text.replace(':', '')
                                    if len(hora_fin_text) >= 3:
                                        if len(hora_fin_text) == 3:
                                            hora_fin_text = '0' + hora_fin_text
                                        h_f = int(hora_fin_text[:2])
                                        m_f = int(hora_fin_text[2:4])
                                        
                                        # Usar fecha de término del evento pero con la hora especificada
                                        f_t_d = Fecha_Termino.replace(hour=h_f, minute=m_f)
                                        fecha['Termina'] = f_t_d.isoformat()
                                    else:
                                        # Si no hay fecha ni hora, usar término del evento
                                        fecha['Termina']=Fecha_Termino.isoformat()
                                except:
                                    fecha['Termina']=Fecha_Termino.isoformat()
                                    
                            
                            
                            
                            periodos[n] = fecha
                            #errores['problemas en invitados'] += 1 if f_n.ids.bg_text == (1,0,0,1) or f_t.ids.bg_text == (1,0,0,1) else 0
                        rut = i.ids.rut_invitado.ids.rut.text
                        errores['problemas en invitados'] += 1 if rut == '' else 0
                        Invitados[rut]=periodos
                        
        if errores['problemas en invitados'] > 0 or errores['campos vacios'] > 0:
            mensaje=''
            mensaje+=f"Campos Vacios o erroneos: {errores['campos vacios']}\n" if errores['campos vacios'] >0 else ''
            mensaje+=f"Problemas en invitados: {errores['problemas en invitados']}" if errores['problemas en invitados'] >0 else ''
            
            emergente = Factory.Alertas_Mensaje()
            emergente.titulo='Error al Crear'
            emergente.texto = mensaje
            emergente.open()
            return
        evento_data={
            'Titulo': titulo,
            'Descripcion': descripcion,
            'Ubicacion': {'Direccion':ubicacion,
                          'Latitud':latitud,
                          'Longitud':longitud
                          },
            'Fechas': {
                'Fecha_Inicio':Fecha_Inicio.isoformat(),
                'Fecha_Termino': Fecha_Termino.isoformat()
                },
            'Estado':'En Espera',
            'Etiquetas': Etiquetas,
            'Archivos':[],
            'Asistencia':[],
            'Comentarios':[],
            'Ganacias':[],
            'HistorialCambios':{
                f"{datetime.now().strftime('%d-%m-%Y')}": {"Cambios":"Se creo y publico el evento"},
            },
            'Organizador': f'{Singleton_Perfil.get_instance().rut}',
            'Acceso': acceso,
            'Visibilidad': visibilidad,
            'Invitados': Invitados
        }
        
        # print(evento_data)
        conn= Escritura_Eventos_DB()
        exito = conn.subir_evento(evento_data)
        
        if exito:
            # Limpiar formulario solo si se creó exitosamente
            self.Limpiar_Campos()
            
            # Mostrar mensaje de éxito
            emergente = Factory.Alertas_Mensaje()
            emergente.titulo = 'Evento Creado'
            emergente.texto = 'El evento se ha creado y publicado exitosamente.'
            emergente.open()
            
            print("✅ Evento creado exitosamente - Formulario limpiado")
        else:
            print("❌ Error al crear evento - No se limpiaron los campos")
            # Mostrar mensaje de error
            emergente = Factory.Alertas_Mensaje()
            emergente.titulo = 'Error al Crear'
            emergente.texto = 'Hubo un problema al subir el evento. Por favor intente nuevamente.'
            emergente.open()
        
        return
        
        
        
        
        
        
        
    
    


